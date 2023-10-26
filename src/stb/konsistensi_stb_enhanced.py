import pandas as pd
import numpy as np
from openpyxl import load_workbook
import yaml
import datetime
import os

base_path = r'C:/Users/amnar/Desktop/gh_konsistensi/'
yaml_path = 'ref/stb_config_all_2021.yml'

# File paths
dataset_path = base_path + 'data/stb/dsB092021STB.xlsx'
reference_files = ['ref/oku.csv',
                   'ref/kewarganegaraan.csv',
                   'ref/kumpulan_etnik.csv', 
                   'ref/persekolahan.csv',
                   'ref/pendidikan_rasmi.csv',
                   'ref/pendidikan_rasmi_tertinggi_2022.csv', 
                   'ref/sijil_tertinggi.csv',
                   'ref/sijil_tertinggi_2022.csv',
                   'ref/status_code.csv', 
                   'ref/msic_code_detail_01.csv',
                   'ref/masco_code.csv',
                   'ref/negara_code.csv', 
                   'ref/institusi_pengajian.csv',
                   'ref/bidang_pengajian.csv']

file_paths = [dataset_path] + [base_path + file for file in reference_files]
yaml_file = base_path + yaml_path

def read_files(*file_paths):
    def read_csv(file_path):
        return pd.read_csv(file_path, encoding='unicode_escape', low_memory=False)

    def read_excel(file_path):
        workbook = load_workbook(filename=file_path)
        sheet_name = workbook.sheetnames[0]  # Get the name of the first sheet
        worksheet = workbook[sheet_name]
        data = list(worksheet.values)
        return pd.DataFrame(data[1:], columns=data[0])

    file_readers = {
        'csv': read_csv,
        'xlsx': read_excel,
        'xls': read_excel
    }
    
    predefined_names = {
        'dsB092021STB.xlsx': 'df',
        'oku.csv': 'df_oku',
        'kewarganegaraan.csv': 'df_kw',
        'kumpulan_etnik.csv': 'df_ket',
        'persekolahan.csv': 'df_persk',
        'pendidikan_rasmi.csv': 'df_pend',
        'pendidikan_rasmi_tertinggi_2022.csv': 'df_pend_22',
        'sijil_tertinggi.csv': 'df_sijil',
        'sijil_tertinggi_2022.csv': 'df_sijil_22',
        'status_code.csv': 'df_status',
        'msic_code_detail_01.csv': 'df_msic',
        'masco_code.csv': 'df_masco',
        'negara_code.csv': 'df_ngra',
        'institusi_pengajian.csv': 'df_ip',
        'bidang_pengajian.csv': 'df_fs'
    }
    
    for file_path in file_paths:
        filename = file_path.split('/')[-1]
        df_name = predefined_names[filename]
        file_format = filename.split('.')[-1]
        
        # Declare the dataframe name as global
        globals()[df_name] = file_readers[file_format](file_path)

# Call the function
read_files(*file_paths)

# Read the YAML file
with open(yaml_file, 'r') as file:
    config = yaml.load(file, Loader=yaml.FullLoader)

# Extract conditions from the YAML data
persekolahan = config['persekolahan']

# Create an empty dictionary to store the merged values
merged_conditions = {}

# Loop over the persekolahan dictionaries and merge the values
for key, conditions in persekolahan.items():
    merged_conditions[key] = {}
    for condition_key, condition_value in conditions.items():
        if condition_key == 'U':
            if isinstance(condition_value, str) and condition_value.isdigit():
                condition_value = int(condition_value)
        merged_conditions[key][condition_key] = condition_value
        
for key in persekolahan:
    persekolahan[key]['U'] = list(eval(persekolahan[key]['U']))

# Extract the dataset name from the file path
dataset_name = os.path.basename(dataset_path)

# Extract the first three characters
doc_type = dataset_name[:3]

# Extract the characters at index 3 and 4
quarter_ref = dataset_name[3:5]

# Extract the characters from index 6 to 9
year_ref = int(dataset_name[5:9])

# Replace the string "None" (and its variants with potential spaces) with NaN for the entire DataFrame
for col in df.columns:
    df[col] = df[col].apply(lambda x: np.nan if str(x).strip() == 'None' else x)

kw_list = df_kw["Kod"].astype(str).apply(lambda x: x.zfill(3)).tolist()
ket_list = list(map(int, df_ket.iloc[:97]["Kod"].values))
status_list = df_status["Kod"].values.tolist()
masco_list = df_masco["KOD_MASCO"].values.tolist()
msic_list = df_msic["KOD_MSIC"].values.tolist()
pkis_list = [str(i).zfill(2) for i in range(1, 13)]
no_kel_list = [str(i).zfill(3) for i in range(1, 1000)]
ngri_list = df_ngra["KOD"].values.tolist()


# -----------------------------------
# Semakan Julat JR4
# -----------------------------------

kel_list = list(range(1000))
b_list = list(range(13))
ng_list = list(range(17))
dp_list = list(range(32))
db_list = list(range(170))
db2_list = set(str(i).zfill(5) for i in range(1, 100000))
st_list = list(range(10))
notk_list = list(range(1000))
noir_list = list(range(100))
t_list = list(range(100))
pkis_list = [str(i).zfill(2) for i in range(1, 13)]
hmis_list = [str(i).zfill(2) for i in range(1, 100)]
bk_list = [str(i).zfill(2) for i in range(1, 13)]
tk_list = [str(i).zfill(4) for i in range(1900, 3000)]
u_list = list(range(201))
tp_list = list(range(6))
ngri_list = [str(i).zfill(2) for i in range(1, 17)] + ['98']
ngra_list = df_ngra["KOD"].astype(str).apply(lambda x: x.zfill(3)).tolist()
oku_list = df_oku["Kod"].astype(str).apply(lambda x: x.zfill(2)).tolist()
pt_22_list = df_pend_22["KOD"].astype(str).apply(lambda x: x.zfill(3)).tolist()
sj_22_list = df_sijil_22["KOD"].astype(str).apply(lambda x: x.zfill(3)).tolist()
ip_list = df_ip["Kod"].tolist()
fs_list = df_fs["Kod"].astype(str).apply(lambda x: x.zfill(4)).tolist()
hmwa_list = list(range(100))

def validate_all_julats(row):
    row['JULAT_001'] = 1 if row['NOKEL'] in no_kel_list else 0
    row['JULAT_002'] = 1 if row['B'] in b_list else 0
    row['JULAT_003'] = 1 if row['NG'] in ng_list else 0
    row['JULAT_004'] = 1 if row['DP'] in dp_list else 0
    row['JULAT_005'] = 1 if row['DB'] in db_list else 0
    row['JULAT_006'] = 1 if row['BP'] in db_list else 0
    row['JULAT_007'] = 1 if row['BP2'] in db2_list else 0
    row['JULAT_008'] = 1 if row['ST'] in st_list else 0
    row['JULAT_009'] = 1 if row['NOTK'] in notk_list else 0
    row['JULAT_010'] = 1 if row['NOIR'] in noir_list else 0
    row['JULAT_011'] = 1 if row['JR'] == 4 else 0
    row['JULAT_012'] = 1 if isinstance(row['NAMA'], str) and len(row['NAMA']) <= 50 else 0
    row['JULAT_013'] = 1 if isinstance(row['NOIC'], str) and len(row['NOIC']) <= 12 or pd.isna(row['NOIC']) else 0
    row['JULAT_014'] = 1 if row['PKIS'] in pkis_list else 0
    row['JULAT_015'] = 1 if row['HMIS'] in hmis_list else 0
    row['JULAT_016'] = 1 if row['J'] in [1, 2] else 0
    row['JULAT_017'] = 1 if row['BK'] in bk_list else 0
    row['JULAT_018'] = 1 if row['TK'] in tk_list else 0
    row['JULAT_019'] = 1 if row['U'] in u_list else 0
    row['JULAT_020'] = 1 if row['KET'] in ket_list else 0
    row['JULAT_021'] = 1 if row['KW'] in kw_list else 0
    row['JULAT_022'] = 1 if row['TP'] in tp_list else 0
    row['JULAT_023'] = 1 if row['NGRI'] in pkis_list else 0
    row['JULAT_024'] = 1 if row['NGRA'] in ngra_list else 0
    row['JULAT_025'] = 1 if row['OKU'] in oku_list else 0
    row['JULAT_026'] = 1 if row['P'] in [1, 2, 3, 4] else 0
    row['JULAT_027'] = 1 if row['PT'] in pt_22_list else 0
    row['JULAT_028'] = 1 if row['SJ'] in sj_22_list else 0
    row['JULAT_029'] = 1 if row['IP'] in ip_list or pd.isna(row['IP']) else 0
    row['JULAT_030'] = 1 if row['FS'] in fs_list else 0
    row['JULAT_031'] = 1 if row['HMWA'] in hmwa_list else 0
    
    return row

df = df.apply(validate_all_julats, axis=1)

# -----------------------------------
# Konsistensi 1(a)
# -----------------------------------

# This function checks a condition based on the 'TK' and 'U' columns of the dataframe.
# If the difference between a reference year and the sum of these columns is greater than 3, 
# it sets the value of 'KONSISTENSI_01a' to 0. Otherwise, it sets the value to 1.
def validate_condition_01(df, year_ref):
    def compute(row):
        if (year_ref - int(row['TK']) - int(row['U'])) > 3:
            return 0
        else:
            return 1
    
    df['KONSISTENSI_01'] = df.apply(lambda row: compute(row) if row['TK'].isdigit() and row['U'].isdigit() else 0, axis=1).astype(int)

validate_condition_01(df, year_ref)

# -----------------------------------
# Konsistensi 1(b)
# -----------------------------------

# This function checks two conditions:
# 1. If 'U' is less than or equal to 15, it sets 'KONSISTENSI_02' to 1.
# 2. If 'U' is greater than 15 and 'HMWA' is not null, it sets 'KONSISTENSI_02' to 1.
def validate_condition_02(df):
    df['KONSISTENSI_02'] = 0
    df.loc[df['U'].astype(int) <= 15, 'KONSISTENSI_02'] = 1
    combined_condition = (df['U'].astype(int) > 15) & ~df['HMWA'].isnull()
    df.loc[combined_condition, 'KONSISTENSI_02'] = 1

validate_condition_02(df)

# -----------------------------------
# Konsistensi 1(c)
# -----------------------------------

# This function checks the values of 'NGRI' for rows where 'NGRA' is 458.
# If 'NGRI' is not in the provided list of allowed values, it sets 'KONSISTENSI_03' to 0.
def validate_condition_03(df, values):
    df['KONSISTENSI_03'] = 1
    df_ngra = df[df['NGRA'] == '458']
    mask = ~df_ngra['NGRI'].isin(values)
    df.loc[df_ngra[mask].index, 'KONSISTENSI_03'] = 0

ngri_lst = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16']

validate_condition_03(df, ngri_lst)

# -----------------------------------
# Konsistensi 1(d)
# -----------------------------------

# This function checks if values in the 'OKU' column are present in the provided list of OKU values.
# If a match is found, it sets 'KONSISTENSI_04' to 1, otherwise 0.
def validate_condition_04(df, oku_list):
    df['KONSISTENSI_04'] = df['OKU'].astype(int).isin(oku_list).astype(int)

oku_list = list(map(int, df_oku["Kod"].values))
validate_condition_04(df, oku_list)

# -----------------------------------
# Konsistensi 1(e)
# -----------------------------------

# This function checks if values in the 'KW' column are present in the provided list.
# If a match is found, it sets 'KONSISTENSI_05a' to 1, otherwise 0.
def validate_condition_05a(df, kw):
    df['KONSISTENSI_05a'] = df['KW'].astype(str).isin(kw).astype(int)

# This function checks if values in the 'KET' column are present in the provided list.
# If a match is found, it sets 'KONSISTENSI_05b' to 1, otherwise 0.
def validate_condition_05b(df, ket):
    df['KONSISTENSI_05b'] = df['KET'].astype(int).isin(ket).astype(int)

validate_condition_05a(df, kw_list)
validate_condition_05b(df, ket_list)

# -----------------------------------
# Konsistensi 1(f)
# -----------------------------------

# This function checks if any row of the dataframe meets any of the conditions specified in merged_conditions.
# If a row meets any of the conditions, it sets 'KONSISTENSI_06' for that row to 1.
def validate_condition_06(df, merged_conditions):
    df['KONSISTENSI_06'] = 0
    for index, row in df.iterrows():
        for condition_key, condition_values in merged_conditions.items():
            match = True
            for col, val in condition_values.items():
                if isinstance(val, list):
                    if row[col] not in val:
                        match = False
                        break
                else:
                    if row[col] != val:
                        match = False
                        break
            if match:
                df.loc[index, 'KONSISTENSI_06'] = 1
                break

validate_condition_06(df, persekolahan)

# -----------------------------------
# Konsistensi 1(g)
# -----------------------------------

# This function checks two conditions:
# 1. If 'U' is greater than or equal to 18, it sets 'KONSISTENSI_08' to 1.
# 2. If 'U' is less than 18 and 'TP' is in the provided list, it sets 'KONSISTENSI_08' to 1.
def validate_condition_08(df, tp_lst):
    df['KONSISTENSI_08'] = 0
    df.loc[df['U'].astype(int) >= 18, 'KONSISTENSI_08'] = 1
    combined_condition = (df['U'].astype(int) < 18) & df['TP'].isin(tp_lst)
    df.loc[combined_condition, 'KONSISTENSI_08'] = 1

validate_condition_08(df, tp_list)

# -----------------------------------
# Konsistensi 1(h)
# -----------------------------------

# This function checks two conditions:
# 1. If 'U' is greater than 15 and 'HMWA' is not null.
# 2. If the first condition is met, it further checks if 'HMWA' and 'HMIS

def validate_condition_09(df):
    df['KONSISTENSI_09'] = 0
    condition_1 = (df['U'].astype(int) > 15) & ~df['HMWA'].isnull()
    condition_2 = df['HMWA'] == df['HMIS']
    df.loc[condition_1 & condition_2, 'KONSISTENSI_09'] = 1

validate_condition_09(df)

# -----------------------------------
# Konsistensi T-C1a
# -----------------------------------

def validate_condition_TC1a(data):
    """
    Validate consistency for 'MSIC_1D', 'S19', and 'S20' in the DataFrame.
    
    Parameters:
        data (pd.DataFrame): DataFrame containing 'MSIC_1D', 'S19', and 'S20'
                             columns to be validated.
        
    Returns:
        pd.DataFrame: DataFrame with an additional column 'KONSISTENSI_TC1a',
                      indicating validation results.
    """
    # Initialize 'KONSISTENSI_TC1a' with 1 for all rows.
    data['KONSISTENSI_TC1a'] = 1
    
    # Identify rows that need further validation based on 'MSIC_1D' and 'S19'.
    filtered_rows = (
        (data['MSIC_1D'] == 'O') &
        data['S19'].between('84111', '84300')
    )
    
    # For the filtered rows, if 'S20' is not '2', mark them as fail (0).
    data.loc[filtered_rows & (data['S20'] != '2'), 'KONSISTENSI_TC1a'] = 0
    
    return data

result = validate_condition_TC1a(df)

# -----------------------------------
# Konsistensi T-C1b
# -----------------------------------

def validate_condition_TC1b(data):
    """
    Validate consistency for 'MSIC_1D', 'S19', and 'KW' in the DataFrame.
    
    Parameters:
        data (pd.DataFrame): DataFrame containing 'MSIC_1D', 'S19', and 'KW'
                             columns to be validated.
        
    Returns:
        pd.DataFrame: DataFrame with an additional column 'KONSISTENSI_TC1b',
                      indicating validation results.
    """
    # Initialize 'KONSISTENSI_TC1b' with 1 for all rows.
    data['KONSISTENSI_TC1b'] = 1
    
    # Identify rows that need further validation based on 'MSIC_1D' and 'S19'.
    filtered_data_TC1b = data[
        (data['MSIC_1D'] == 'O') &
        data['S19'].between('84111', '84300')
    ].copy()
    
    # For the filtered rows, if 'KW' is 458, mark them as fail (0).
    data.loc[data.index.isin(filtered_data_TC1b.index) & (data['KW'] == 458), 'KONSISTENSI_TC1b'] = 0
    
    return data

result = validate_condition_TC1b(df)

# -----------------------------------
# Konsistensi T-C2
# -----------------------------------

def validate_condition_TC2(data):
    """
    Validate consistency based on conditions for 'MSIC_1D' and 'S20'.
    
    Parameters:
        data (pd.DataFrame): DataFrame containing 'MSIC_1D' and 'S20'
                             columns to be validated.
        
    Returns:
        pd.DataFrame: DataFrame with an additional column 'KONSISTENSI_TC2',
                      indicating validation results.
    """
    # Initialize 'KONSISTENSI_TC2' with 1 for all rows (assume pass initially).
    data['KONSISTENSI_TC2'] = 1
    
    # Identify rows that need further validation based on 'MSIC_1D'.
    filtered_data_TC2 = data[(data['MSIC_1D'] == 'T')].copy()
    
    # For the filtered rows, if 'S20' is not 3 and not NaN, mark them as fail (0).
    data.loc[data.index.isin(filtered_data_TC2.index) & (data['S20'].astype(str) != 3) & (data['S20'].notna()), 'KONSISTENSI_TC2'] = 0      
    
    return data
    
result = validate_condition_TC2(df)

# -----------------------------------
# Konsistensi T-C3
# -----------------------------------

def validate_condition_TC3(data):
    """
    Validate consistency based on conditions for 'MSIC_1D' and 'S20'.
    
    Parameters:
        data (pd.DataFrame): DataFrame containing 'MSIC_1D' and 'S20'
                             columns to be validated.
        
    Returns:
        pd.DataFrame: DataFrame with an additional column 'KONSISTENSI_TC3',
                      indicating validation results.
    """
    # Initialize 'KONSISTENSI_TC3' with 1 for all rows (assume pass initially).
    data['KONSISTENSI_TC3'] = 1
    
    # Identify rows that need further validation based on 'MSIC_1D'.
    filtered_data_TC3 = data[(data['MSIC_1D'] == 'P')].copy()
    
    # For the filtered rows, if 'S20' is not 2 and not NaN, mark them as fail (0).
    data.loc[data.index.isin(filtered_data_TC3.index) & (data['S20'].astype(str) != '2') & (data['S20'].notna()), 'KONSISTENSI_TC3'] = 0             
    
    return data

result = validate_condition_TC3(df)

# -----------------------------------
# Konsistensi T-C4
# -----------------------------------

def validate_condition_TC4(data):
    """
    Validate consistency based on conditions for 'MSIC_1D', 'MASCO_1D', and 'S20'.
    
    Parameters:
        data (pd.DataFrame): DataFrame containing 'MSIC_1D', 'MASCO_1D', and 'S20' 
                             columns to be validated.
        
    Returns:
        pd.DataFrame: DataFrame with an additional column 'KONSISTENSI_TC4',
                      indicating validation results.
    """
    # Initialize 'KONSISTENSI_TC4' with 1 for all rows (assume pass initially).
    data['KONSISTENSI_TC4'] = 1
    
    # Identify rows that need further validation based on 'MSIC_1D' and 'MASCO_1D'.
    filtered_data_TC4 = data[(data['MSIC_1D'] == 'A') & (data['MASCO_1D'] == 9)].copy()
    
    # For the filtered rows, if 'S20' is not 3, 4, or 5 and not NaN, mark them as fail (0).
    data.loc[data.index.isin(filtered_data_TC4.index) & (~data['S20'].isin(['3', '4', '5'])) & (data['S20'].notna()), 'KONSISTENSI_TC4'] = 0       
    
    return data

result = validate_condition_TC4(df)

# -----------------------------------
# Konsistensi T-C5a
# -----------------------------------

def validate_condition_TC5a(data):
    """
    Validate consistency based on conditions for 'S18' and 'SJ'.
    
    Parameters:
        data (pd.DataFrame): DataFrame containing 'S18' and 'SJ' columns to be validated.
        
    Returns:
        pd.DataFrame: DataFrame with an additional column 'KONSISTENSI_TC5a',
                      indicating validation results.
    """
    # Ensure 'SJ' is treated as a numeric column, converting invalid entries to NaN.
    data['SJ'] = pd.to_numeric(data['SJ'], errors='coerce')
    
    # Initialize 'KONSISTENSI_TC5a' with 1 for all rows (assume pass initially).
    data['KONSISTENSI_TC5a'] = 1
    
    # Identify rows that meet the first condition for further validation.
    filtered_data_TC5a = data[(data['S18'] >= '111101') & (data['S18'] <= '291918')].copy()
    
    # For the filtered rows, if 'SJ' is NOT between 20 and 242 and is not NaN, 
    # mark them as fail (0).
    data.loc[data.index.isin(filtered_data_TC5a.index) & (~data['SJ'].between(20, 242)) & (data['SJ'].notna()), 'KONSISTENSI_TC5a'] = 0       
    
    return data

result = validate_condition_TC5a(df)

# -----------------------------------
# Konsistensi T-C5b
# -----------------------------------

def validate_condition_TC5b(data):
    """
    Validate consistency based on conditions for 'S18' and 'S20'.
    
    Conditions:
    1. Rows where 'S18' is between '111101' and '291918' are considered for validation.
    2. Initially, all entries are assumed to pass (KONSISTENSI_TC5b = 1).
    3. Entries that fulfill condition 1 but have 'S20' equal to '4' are marked as fail (KONSISTENSI_TC5b = 0).
    
    Parameters:
        data (pd.DataFrame): DataFrame containing 'S18' and 'S20' columns to be validated.
        
    Returns:
        pd.DataFrame: DataFrame with an additional column 'KONSISTENSI_TC5b',
                      indicating validation results.
    """
    # Ensure 'S18' and 'S20' are treated as strings.
    data[['S18', 'S20']] = data[['S18', 'S20']].astype(str)
    
    # Initialize 'KONSISTENSI_TC5b' with 1 for all rows (assume pass initially).
    data['KONSISTENSI_TC5b'] = 1
    
    # Identify rows that meet the first condition for further validation.
    filtered_data_TC5b = data[data['S18'].between('111101', '291918')].copy()
    
    # For the filtered rows, if 'S20' is '4', mark them as fail (0).
    data.loc[data.index.isin(filtered_data_TC5b.index) & (data['S20'] == '4'), 'KONSISTENSI_TC5b'] = 0
    
    return data

result = validate_condition_TC5b(df)

# -----------------------------------
# Konsistensi T-C6
# -----------------------------------

def validate_condition_TC6(data):
    """
    Validate consistency based on conditions for 'STATUS' and 'S20'.
    
    Conditions:
    1. Rows where 'S20' is equal to 2 are considered for validation.
    2. Initially, all entries are assumed to pass (KONSISTENSI_TC6 = 1).
    3. Entries that fulfill condition 1 but have 'STATUS' not equal to 'GOV' and not NaN are marked as fail (KONSISTENSI_TC6 = 0).
    
    Parameters:
        data (pd.DataFrame): DataFrame containing 'STATUS' and 'S20' columns to be validated.
        
    Returns:
        pd.DataFrame: DataFrame with an additional column 'KONSISTENSI_TC6',
                      indicating validation results.
    """
    # Ensure 'S20' is treated as a string. 'STATUS' will also be treated as string but NaN will be kept as is.
    data['S20'] = data['S20'].astype(str)
    
    # Initialize 'KONSISTENSI_TC6' with 1 for all rows (assume pass initially).
    data['KONSISTENSI_TC6'] = 1
    
    # Identify rows that meet the first condition for further validation.
    filtered_data_TC6 = data[data['S20'] == '2'].copy()
    
    # For the filtered rows, if 'STATUS' is not 'GOV' and not NaN, mark them as fail (0).
    data.loc[data.index.isin(filtered_data_TC6.index) & (data['STATUS'] != 'GOV') & data['STATUS'].notna(), 'KONSISTENSI_TC6'] = 0      
    
    return data

result = validate_condition_TC6(df)

# -----------------------------------
# Konsistensi T-C7
# -----------------------------------

def validate_condition_TC7(data, masco_list):
    # Initialize a new column 'KONSISTENSI_TC7' and set it as 1 (pass) for all rows initially
    data['KONSISTENSI_TC7'] = 1
    
    # Filter data according to the first condition
    filtered_data_TC7 = data[(data['S18'].isin(masco_list)) & (data['S20'] == '2')].copy()
    
    # Apply the second condition:
    # For the filtered rows, if 'STATUS' is not 'GOV' and is not NaN, mark them as fail (0).
    data.loc[data.index.isin(filtered_data_TC7.index) & (data['STATUS'] != 'GOV') & data['STATUS'].notna(), 'KONSISTENSI_TC7'] = 0    
    
    return data

result = validate_condition_TC7(df, masco_list)

# -----------------------------------
# Konsistensi T-C8a
# -----------------------------------

def validate_condition_TC8a(data):
    """
    Validates the condition T-C8a.

    Parameters:
        data (pd.DataFrame): The data on which to perform the validation.

    Returns:
        pd.DataFrame: The data with an additional column indicating the 
                      results of the validation.
                      
    Explanation:
        - First Condition: 
            The data is filtered based on the condition (data['S20'] == 2). 
            Rows that do not satisfy this condition are considered as pass 
            (KONSISTENSI_TC8a = 1).
            
        - Second Condition: 
            From the filtered data, the 'S19' should be between '84111' and '84300' 
            (inclusive). If 'S19' is outside of this range, it's considered as fail 
            (KONSISTENSI_TC8a = 0). Otherwise, it's considered as pass (KONSISTENSI_TC8a = 1).
    """
    
    # Initialize the new column with 1 (indicating pass) for all rows.
    data['KONSISTENSI_TC8a'] = 1

    # Identify the rows that meet the first condition: 'S20' should be equal to 2.
    filtered_data_TC8a = data[(data['S20'] == 2)].copy()

    # For the filtered rows, if 'S19' is outside the range '84111' to '84300', mark them as fail (KONSISTENSI_TC8a = 0).
    data.loc[data.index.isin(filtered_data_TC8a.index) & (~data['S19'].astype(str).between('84111', '84300')), 'KONSISTENSI_TC8a'] = 0     
    
    return data

result = validate_condition_TC8a(df)

# -----------------------------------
# Konsistensi T-C8b
# -----------------------------------

def validate_condition_TC8b(data):
    """
    Validates the condition T-C8b.

    Parameters:
        data (pd.DataFrame): The data on which to perform the validation.

    Returns:
        pd.DataFrame: The data with an additional column indicating the 
                      results of the validation.
                      
    Explanation:
        - First Condition: 
            The data is filtered based on the condition (data['S20'] == 4). 
            Rows that do not satisfy this condition are considered as pass 
            (KONSISTENSI_TC8b = 1).
            
        - Second Condition: 
            From the filtered data, if 'S19' is NaN (i.e., does not have a value), 
            it's considered as fail (KONSISTENSI_TC8b = 0). Otherwise, it's considered 
            as pass (KONSISTENSI_TC8b = 1).
    """
    
    # Initialize the new column with 1 (indicating pass) for all rows.
    data['KONSISTENSI_TC8b'] = 1

    # Identify the rows that meet the first condition: 'S20' should be equal to 4.
    filtered_data_TC8b = data[(data['S20'] == 4)].copy()

    # For the filtered rows, if 'S19' is NaN, mark them as fail (KONSISTENSI_TC8b = 0).
    data.loc[data.index.isin(filtered_data_TC8b.index) & data['S19'].isna(), 'KONSISTENSI_TC8b'] = 0
    
    return data

result = validate_condition_TC8b(df)

# -----------------------------------
# Konsistensi T-C9
# -----------------------------------

def validate_condition_TC9(data):
    """
    This function performs the T-C9 validation on the dataframe `data`.
    
    Parameters:
        - data (pd.DataFrame): the input data to validate.
        
    Returns:
        pd.DataFrame: the input dataframe with an additional column `KONSISTENSI_TC9`
                      which indicates pass (1) or fail (0) for the validation.
    
    The validation conditions are as follows:
        1. Filter rows with PKIS == 11. The rest are considered pass.
        2. For the filtered rows from condition 1, if MSIC_1D is not 'T' and not NaN, mark them as fail (0).
           If MSIC_1D is 'T' or NaN, they pass this validation stage (1).
    """
    # Create a new column 'KONSISTENSI_TC9' and initialize it with 1 (pass)
    data['KONSISTENSI_TC9'] = 1
    
    # First condition: Filter rows where 'PKIS' is 11
    filtered_data_TC9 = data[data['PKIS'].astype(str) == '11'].copy()
    
    # Second condition:
    # For the filtered rows, if 'MSIC_1D' is not 'T' and not NaN, mark them as fail (0).
    data.loc[data.index.isin(filtered_data_TC9.index) & (data['MSIC_1D'] != 'T') & (data['MSIC_1D'].notna()), 'KONSISTENSI_TC9'] = 0          
    
    return data

result = validate_condition_TC9(df)

# -----------------------------------
# Konsistensi T-C10
# -----------------------------------

def validate_condition_TC10(data):
    """
    Validate the condition TC10 on the provided dataframe.

    Condition:
    1. First, filter the data where 'S19' is between '98100' and '98200'.
    2. For the filtered data, if 'RIN' is NaN, it is considered pass (KONSISTENSI_TC10 = 1).
       Else, it is considered fail (KONSISTENSI_TC10 = 0).

    Parameters:
        data (pd.DataFrame): The dataframe to validate.

    Returns:
        pd.DataFrame: The dataframe with the validation results.
    """
    
    # Ensure the 'S19' column is a string, so we can perform string comparisons
    data['S19'] = data['S19'].astype(str)
    
    # First condition: Filter data where 'S19' is between '98100' and '98200'
    filtered_data_TC10 = data[data['S19'].between('98100', '98200')].copy()
    
    # Initialize a new column for the validation results and set to 1 as default (pass)
    data['KONSISTENSI_TC10'] = 1
    
    # Second condition: For the filtered rows, if 'RIN' is not NaN, mark them as fail (0)
    data.loc[data.index.isin(filtered_data_TC10.index) & data['RIN'].notna(), 'KONSISTENSI_TC10'] = 0
    
    return data

result = validate_condition_TC10(df)

# -----------------------------------
# Konsistensi T-C11
# -----------------------------------

def validate_condition_TC11(data):
    """
    Validate conditions for T-C11.
    
    The data is first filtered based on 'RIN' being between 1 and 5. Rows that do not
    meet this condition are considered as 'pass' (KONSISTENSI_TC11 = 1).
    
    From the remaining filtered data, the function checks whether 'S19' equals '99000'.
    If 'S19' is not '99000' or is NaN, it marks them as 'fail' (KONSISTENSI_TC11 = 0).
    """
    
    # Ensure the 'RIN' and 'S19' columns are treated as strings for consistency in comparison
    data['RIN'] = data['RIN'].astype(str)
    data['S19'] = data['S19'].astype(str)
    
    # First condition: filter rows where 'RIN' is between '1' and '5'.
    filtered_data_TC11 = data[data['RIN'].isin(['1', '2', '3', '4', '5'])].copy()
    
    # Initialize the validation column with 'pass' (1).
    data['KONSISTENSI_TC11'] = 1
    
    # Second condition:
    # For the filtered rows, if 'S19' is not '99000' and not NaN, mark them as 'fail' (0).
    data.loc[data.index.isin(filtered_data_TC11.index) & (~data['S19'].isin(['99000'])) & (data['S19'].notna()), 'KONSISTENSI_TC11'] = 0      
    
    return data

result = validate_condition_TC11(df)

# -----------------------------------
# Konsistensi T-C12
# -----------------------------------

def validate_condition_TC12(data):
    """
    Validate the consistency of the data against conditions (T-C12).
    
    This function checks the data for the following consistency conditions:
    - First, it filters the data for rows where 'MASCO_1D' and 'MSIC_1D' have non-NaN values.
    - Then, it checks whether 'S19' is between '98100' and '98200' (inclusive) for the filtered data.
      If 'S19' is within this range, the data does not meet the consistency requirement.
    
    Parameters:
    - data (pd.DataFrame): The data on which to perform the consistency check.
    
    Returns:
    pd.DataFrame: The original data with an additional column 'KONSISTENSI_TC12',
                  indicating pass (1) or fail (0) for each row based on the consistency check.
    """
    
    # First condition: 
    # Filter rows where both 'MASCO_1D' and 'MSIC_1D' have valid (non-NaN) values.
    filtered_data_TC12 = data[data['MASCO_1D'].notna() & data['MSIC_1D'].notna()].copy()
    
    # Initialize a new column 'KONSISTENSI_TC12' with 1 (indicating pass) for all rows.
    data['KONSISTENSI_TC12'] = 1

    # Second condition: 
    # For the filtered rows, if 'S19' is within the range ['98100', '98200'], mark them as fail (0).
    data.loc[data.index.isin(filtered_data_TC12.index) & data['S19'].between('98100', '98200'), 'KONSISTENSI_TC12'] = 0        

    return data

result = validate_condition_TC12(df)

# -----------------------------------
# Konsistensi T-C13
# -----------------------------------

def validate_condition_TC13(data):
    """
    Validate the data according to the T-C13 condition.
    
    Conditions:
    1. First, filter the data where 'KW' is equal to 458. 
       If 'KW' is not equal to 458, the entry is considered as pass (KONSISTENSI_TC13 = 1).
    2. From the filtered data, check if the 'KET' value is in the provided `ket_list`.
       If 'KET' is NOT in `ket_list` or is NaN, it's considered a fail (KONSISTENSI_TC13 = 0).
       Otherwise, it's considered a pass (KONSISTENSI_TC13 = 1).
       
    Parameters:
    - data (DataFrame): The data on which to perform the validation.
    - ket_list (list): A list of valid values for the 'KET' field.
    
    Returns:
    DataFrame: The data with an additional column 'KONSISTENSI_TC13' indicating pass/fail.
    """
    # Ensure 'KW' and 'KET' are treated as strings
    data['KW'] = data['KW'].astype(str)
    data['KET'] = data['KET'].astype(str)
    
    # Initialize a new column for the validation results and set default as 1 (pass)
    data['KONSISTENSI_TC13'] = 1
    
    # Identify rows that meet the first condition
    condition_1 = (data['KW'] == '458')
    
    # For the rows meeting the first condition, if 'KET' is NOT in ket_list or is NaN, mark them as fail (0)
    data.loc[condition_1 & (~data['KET'].isin(ket_list) | data['KET'].isna()), 'KONSISTENSI_TC13'] = 0
    
    return data

result = validate_condition_TC13(df)

# -----------------------------------
# Konsistensi T-C14
# -----------------------------------

# TC-14: Jika Jantina (J) Lelaki (kod 1), 
#     check PKIS = 01 (Ketua Isi Rumah) dan S10 = 02 (kerja rumah/ tanggungjawab keluarga). 
#     Jika Umur (U) >= 50 boleh terima, tapi jika U < 50, semak semula

def validate_condition_TC14(data):
    """
    Validate the condition TC-14:
    - First, filter rows where 'J' == '1', 'PKIS' == '01', and 'S10' == '2' (first condition).
    - For the filtered rows, if 'U' < 50, mark them as fail (0).
    """
    # Ensure the columns are in the correct data type
    data['J'] = data['J'].astype(str)
    data['PKIS'] = data['PKIS'].astype(str)
    data['S10'] = data['S10'].astype(str)
    data['U'] = pd.to_numeric(data['U'], errors='coerce')
    
    # Initialize a new column for the validation result and set default as 1 (pass)
    data['KONSISTENSI_TC14'] = 1

    # Filter the data to meet the first condition
    filtered_data_TC14 = data[(data['J'] == '1') & (data['PKIS'] == '01') & (data['S10'] == '2')].copy()

    # For the filtered rows, if 'U' < 50, mark them as fail (0)
    data.loc[data.index.isin(filtered_data_TC14.index) & (data['U'] < 50), 'KONSISTENSI_TC14'] = 0

    return data

result = validate_condition_TC14(df)

# -----------------------------------
# Konsistensi T-C15a & T-C15b
# -----------------------------------

def validate_condition_TC15(data):
    # Ensure the U column is treated as numeric
    data['U'] = pd.to_numeric(data['U'], errors='coerce')
    
    # Initializing columns with 1 (pass)
    data['KONSISTENSI_TC15a'] = 1
    data['KONSISTENSI_TC15b'] = 1
    
    # First condition: T-C15a
    # Filter where TP == 1 and PKIS is in ['02', '04', '05', '07', '08']
    condition_TC15a = (data['TP'] == 1) & data['PKIS'].isin(['02', '04', '05', '07', '08'])
    # Mark as fail (0) if the condition is met
    data.loc[condition_TC15a, 'KONSISTENSI_TC15a'] = 0
    
    # Second condition: T-C15b
    # Filter where TP in [2, 3, 4, 5] and U <= 17
    condition_TC15b = data['TP'].isin([2, 3, 4, 5]) & (data['U'] <= 17)
    # Mark as fail (0) if the condition is met
    data.loc[condition_TC15b, 'KONSISTENSI_TC15b'] = 0
    
    # Combined condition: T-C15
    # Initialize with 1 (pass)
    data['KONSISTENSI_TC15'] = 1
    # If either of the sub-conditions (T-C15a or T-C15b) is 0 (fail), mark T-C15 as 0 (fail)
    data.loc[(data['KONSISTENSI_TC15a'] == 0) | (data['KONSISTENSI_TC15b'] == 0), 'KONSISTENSI_TC15'] = 0
    
    return data


result = validate_condition_TC15(df)

# -----------------------------------
# Konsistensi T-C16a
# -----------------------------------

def validate_condition_TC16a(data):
    """
    Validate the consistency condition TC16a.
    
    Conditions:
    - P = 1 (Not schooling)
    - S10 = 01 and S15 = 1 OR S10 = 07 and S15 in [2, 3]
    
    If a row satisfies both conditions, it will be marked as fail (0).
    Otherwise, it will be marked as pass (1).
    
    Parameters:
        data (DataFrame): The data to validate.
    
    Returns:
        DataFrame: The data with a new column 'KONSISTENSI_TC16a' indicating pass/fail.
    """
    # Filtering the data according to the specified conditions
    filtered_data_TC16a = data[
        (data['P'] == 1) &
        (
            ((data['S10'] == '01') & (data['S15'] == '1')) |
            ((data['S10'] == '07') & data['S15'].isin(['2', '3']))
        )
    ].copy()
    
    # Initializing a new column with 1 (pass)
    data['KONSISTENSI_TC16a'] = 1
    
    # Marking the rows that fail the conditions as 0 (fail)
    data.loc[filtered_data_TC16a.index, 'KONSISTENSI_TC16a'] = 0

    return data

result = validate_condition_TC16a(df)

# -----------------------------------
# Konsistensi T-C16b
# -----------------------------------

def validate_condition_TC16b(data):
    """
    Validate the consistency condition TC16b.
    
    Conditions:
    - P = 2 (Schooling)
    - S10 in [03, 08, 09, 12, 13]
    
    If a row satisfies both conditions, it will be marked as fail (0).
    Otherwise, it will be marked as pass (1).
    
    Parameters:
        data (DataFrame): The data to validate.
    
    Returns:
        DataFrame: The data with a new column 'KONSISTENSI_TC16b' indicating pass/fail.
    """
    # Filtering the data according to the specified conditions
    filtered_data_TC16b = data[
        (data['P'] == 2) & 
        data['S10'].isin(['03', '08', '09', '12', '13'])
    ].copy()
    
    # Initializing a new column with 1 (pass)
    data['KONSISTENSI_TC16b'] = 1
    
    # Marking the rows that fail the conditions as 0 (fail)
    data.loc[filtered_data_TC16b.index, 'KONSISTENSI_TC16b'] = 0

    return data

result = validate_condition_TC16b(df)
                               
# -----------------------------------
# Konsistensi T-C16c
# -----------------------------------

def validate_condition_TC16c(data):
    """
    Validate the consistency condition TC16c.
    
    Conditions:
    - P = 4 (Finished school)
    - S10 in [01, 07]
    
    If a row satisfies both conditions, it will be marked as fail (0).
    Otherwise, it will be marked as pass (1).
    
    Parameters:
        data (DataFrame): The data to validate.
    
    Returns:
        DataFrame: The data with a new column 'KONSISTENSI_TC16c' indicating pass/fail.
    """
    # Filtering the data according to the specified conditions
    filtered_data_TC16c = data[
        (data['P'] == 4) & 
        data['S10'].isin(['01', '07'])
    ].copy()
    
    # Initializing a new column with 1 (pass)
    data['KONSISTENSI_TC16c'] = 1
    
    # Marking the rows that fail the conditions as 0 (fail)
    data.loc[filtered_data_TC16c.index, 'KONSISTENSI_TC16c'] = 0

    return data

result = validate_condition_TC16c(df)

# Extract the original file name
original_file_name = os.path.basename(dataset_path)

suffix = '_konsistensi'

# Create the new file name by adding the suffix and changing the extension to .xlsx
new_file_name = original_file_name.replace('.xlsx', '') + suffix + '.xlsx'

# Save the DataFrame to Excel using the new file name
df.to_excel(os.path.join(output_file_path, new_file_name), index=False)

# Extract the original file name
original_file_name = os.path.basename(dataset_path)

suffix = '_konsistensi'

# Create the new file name by adding the suffix
new_file_name = original_file_name.replace('.csv', '') + suffix + '.csv'

# Save the DataFrame as CSV using the new file name
df.to_csv(os.path.join(output_file_path, new_file_name), index=False)
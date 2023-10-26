import pandas as pd
import numpy as np
from openpyxl import load_workbook
import re
import os
import glob

base_path = r'C:/Users/amnar/Desktop/gh_konsistensi/'

# File paths
# present_path = base_path + 'data/sgtgu/2022_RawData QESWS 3Q.xlsb'
# past_file = base_path + 'data/sgtgu/2022_RawData QESWS 2Q.xlsb'
dataset_path = base_path + 'data/sgtgu/2022_RawData QESWS 3Q_500_new_01.xlsx'

reference_files = ['ref/msic_code_detail_01.csv',
                   'ref/survey_code.csv']

file_paths = [dataset_path] + [base_path + file for file in reference_files]

def read_files(*file_paths):
    def read_csv(file_path):
        return pd.read_csv(file_path, encoding='unicode_escape', low_memory=False)

    def read_excel(file_path):
        workbook = load_workbook(filename=file_path)
        sheet_name = workbook.sheetnames[0]  # Get the name of the first sheet
        worksheet = workbook[sheet_name]
        data = list(worksheet.values)
        return pd.DataFrame(data[1:], columns=data[0])

    file_readers = {
        'csv': read_csv,
        'xlsx': read_excel,
        'xls': read_excel
    }
    
    predefined_names = {
#         '2022_RawData QESWS 3Q.xlsb': 'df_present',
#         '2022_RawData QESWS 2Q.xlsb': 'df_past',
        '2022_RawData QESWS 3Q_500_new_01.xlsx': 'df',
        'msic_code_detail_01.csv': 'df_msic',
        'survey_code.csv': 'df_survey'
    }
    
    for file_path in file_paths:
        filename = file_path.split('/')[-1]
        df_name = predefined_names[filename]
        file_format = filename.split('.')[-1]
        
        # Declare the dataframe name as global
        globals()[df_name] = file_readers[file_format](file_path)

# Call the function
read_files(*file_paths)

print(df.shape)

# Input for year and quarter
year = 22
quarter = 3

df_new = df[(df['STATUS_present'] == 4)]
df_new.shape

# -----------------------------------
# Query 1
# -----------------------------------

"""
This function is applied row-wise to a DataFrame to validate the consistency and integrity of several columns, namely, 'RESPONSE_CODE_past', 'RESPONSE_CODE_present', 'MSIC2008_SMPL_present', 'MSIC2008_present', 'MSIC2008_past', 'RECEIPT_MODE_present', 'SURVEY_CODE_present', and 'MSIC2008_SMPL_present'.

Logic:
1. The values in 'RESPONSE_CODE_past' and 'RESPONSE_CODE_present' should be the same.
2. The values in 'MSIC2008_SMPL_present' and 'MSIC2008_present' should be the same.
3. The values in 'MSIC2008_present' and 'MSIC2008_past' should be the same.
4. 'RESPONSE_CODE_present' should be within the predefined 'resp_code' list.
5. 'RECEIPT_MODE_present' should be within the predefined 'reci_code' list.
6. 'SURVEY_CODE_present' should be within the predefined 'survey_code' list.
7. 'MSIC2008_SMPL_present' should be within the predefined 'msic_code' list.
8. 'MSIC2008_present' should not be NaN or an empty string.

The results of this validation are stored in the 'QUERY_01' column, where each value indicates whether the condition holds (1 for True, 0 for False).

Parameters:
- row : pandas Series
    The current row of the DataFrame being processed.

Returns:
- int
    An integer (0 or 1) indicating the result of the validation for the row.
"""

# Create list based on msic code and survey code
msic_code = df_msic["KOD_MSIC"].values.astype(str).tolist()
survey_code = df_survey["KOD"].values.astype(str).tolist()

# Response code and receipt mode
resp_code = [11, 12, 13, 14, 15, 21, 22, 23, 31, 32, 40, 50, 60, 71, 72, 73, 74, 75, 76, 77]
reci_code = [1, 2, 3, 4, 5, 6]

# Main function
def query_01(row):
    # Check if RESPONSE_CODE_past and RESPONSE_CODE_present are not the same
    if row['RESPONSE_CODE_past'] != row['RESPONSE_CODE_present']:
        return 0
    
    # Check if MSIC2008_SMPL_present and MSIC2008_SMPL_past are not the same
    if row['MSIC2008_SMPL_present'] != row['MSIC2008_present']:
        return 0
    
    # Check if MSIC2008_present and MSIC2008_past are not the same
    if row['MSIC2008_present'] != row['MSIC2008_past']:
        return 0
    
    if (row['RESPONSE_CODE_present'] in resp_code) + \
       (row['RECEIPT_MODE_present'] in reci_code) + \
       (row['SURVEY_CODE_present'] in survey_code) + \
       (row['MSIC2008_SMPL_present'] in msic_code) + \
       ~(pd.isna(row['MSIC2008_present']) or row['MSIC2008_present'] == ''):
        return 1
    
    return 0
    
df['QUERY_01'] = df.apply(query_01, axis = 1)

# -----------------------------------
# Query 2
# -----------------------------------

def generate_lists_02(year, quarter,
                      prefix_01='_F0110',
                      prefix_02='_F0710',
                      prefix_03='_F1310'):

    def get_month_quarter(quarter):
        if quarter == 1:
            return ['01', '02', '03']
        elif quarter == 2:
            return ['04', '05', '06']
        elif quarter == 3:
            return ['07', '08', '09']
        elif quarter == 4:
            return ['10', '11', '12']
        else:
            return ['No quarter found']

    month_quarter = get_month_quarter(quarter)
    
    col_01 = f'A{year}{quarter}{month_quarter[0]}{prefix_01}'
    col_02 = f'A{year}{quarter}{month_quarter[1]}{prefix_02}'
    col_03 = f'A{year}{quarter}{month_quarter[2]}{prefix_03}'
    
    return col_01, col_02, col_03

col_01, col_02, col_03 = generate_lists_02(year, quarter)

def query_02(df, col_01, col_02, col_03):

    """
    This function is applied to a DataFrame to validate specific conditions concerning three columns (col_01, col_02, col_03) in relation to the 'RESPONSE_CODE_present' and 'IND_STATUS_past' columns.

    Logic:
    1. Identify rows where 'RESPONSE_CODE_present' is equal to 11 and either 'IND_STATUS_past' is NaN or 'IND_STATUS_past' is 0.
    2. For the identified rows:
        - If any of the values in the three columns (col_01, col_02, col_03) is 0 or NaN, the validation fails for that row.

    The results of this validation are stored in a new column named 'QUERY_02', where:
        - 1 indicates a pass (conditions hold true)
        - 0 indicates a fail (conditions do not hold true)

    Parameters:
    - df : pandas DataFrame
        The main DataFrame containing the data.
    - col_01, col_02, col_03 : str
        Names of the columns to be validated.

    Returns:
    - pandas DataFrame
        Updated DataFrame with the 'QUERY_02' column added containing the results of the validation.
    """
    
    # Define check_columns list
    check_columns = [col_01, col_02, col_03]
    
    # Initialize 'QUERY_02' column with 1 (consider all rows as pass by default)
    df['QUERY_02'] = 1
    
    # Filter the dataframe based on the primary condition to identify potential rows to be flagged as 'Fail'
    filtered_df = df[(df['RESPONSE_CODE_present'] == 11) & (df['IND_STATUS_past'].isna() | (df['IND_STATUS_past'] == 0))]
    
    # Check for zero or NaN values in the specified columns among the filtered rows
    zero_or_nan_value_rows = filtered_df[check_columns].eq(0).any(axis=1) | filtered_df[check_columns].isna().any(axis=1)
    
    # Update 'QUERY_02' column based on the sub-conditions for the filtered rows
    # Rows with any zero or NaN value in the specified columns among the filtered rows are flagged as 'Fail' (0)
    df.loc[filtered_df[zero_or_nan_value_rows].index, 'QUERY_02'] = 0
    
    return df

df = query_02(df, col_01, col_02, col_03)

# Filter data with this condition for query 3 - 11
df = df[(df['RESPONSE_CODE_present'] == 11) & (df['RESPONSE_CODE_past'] == 11)]
df.shape

# -----------------------------------
# Query 3
# -----------------------------------

def generate_lists_03(df, quarter, year,
                      A1_past_prefix=None, A1_present_prefix=None,
                      A2_present_prefix=None, A3_present_prefix=None):

    def get_month_quarter(quarter):
        if quarter == 1:
            return ['01', '02', '03']
        elif quarter == 2:
            return ['04', '05', '06']
        elif quarter == 3:
            return ['07', '08', '09']
        elif quarter == 4:
            return ['10', '11', '12']
        else:
            return ['No quarter found']

    month_quarter = get_month_quarter(quarter)
    past_quarter = quarter - 1  # Calculate previous quarter

    if past_quarter < 1:
        past_quarter = 4  # Wrap around to the fourth quarter if current quarter is 1

    past_month_quarter = get_month_quarter(past_quarter)  # Get month range for previous quarter

    if A1_past_prefix is None:
        A1_past_prefix = past_month_quarter[-1]  # Set A1_past_prefix to the last month of previous quarter
    if A1_present_prefix is None:
        A1_present_prefix = month_quarter[0]
    if A2_present_prefix is None:
        A2_present_prefix = month_quarter[1]
    if A3_present_prefix is None:
        A3_present_prefix = month_quarter[2]

    A1_past_lst = []
    A1_present_lst = []
    A2_present_lst = []
    A3_present_lst = []

    for i in range(1, 11):
        # Subtract 1 from the year only if the quarter is 1
        past_year = year - 1 if quarter == 1 else year
        A1_past_col = f'A{past_year}{past_quarter}{A1_past_prefix}_F13{i:02d}'
        A1_present_col = f'A{year}{quarter}{A1_present_prefix}_F01{i:02d}'
        A2_present_col = f'A{year}{quarter}{A2_present_prefix}_F07{i:02d}'
        A3_present_col = f'A{year}{quarter}{A3_present_prefix}_F13{i:02d}'

        A1_past_lst.append(A1_past_col)
        A1_present_lst.append(A1_present_col)
        A2_present_lst.append(A2_present_col)
        A3_present_lst.append(A3_present_col)

    return A1_past_lst, A1_present_lst, A2_present_lst, A3_present_lst, month_quarter

lists_tuple = generate_lists_03(df, quarter, year)
A1_past_lst, A1_present_lst, A2_present_lst, A3_present_lst, month_quarter = lists_tuple

column_pairs_01 = [(a2, a1) for a1, a2 in zip(A1_past_lst, A1_present_lst)]
column_pairs_02 = [(a2, a1) for a1, a2 in zip(A1_present_lst, A2_present_lst)]
column_pairs_03 = [(a2, a1) for a1, a2 in zip(A2_present_lst, A3_present_lst)]

def query_03(df, column_pairs):
    
    """
    This function is applied to a DataFrame to validate the growth percentage between two columns. 
    The growth is computed between each pair of columns provided in `column_pairs`.

    Logic:
    1. Calculate the growth percentage using the formula: ((column1 / column2) - 1) * 100
    2. If the growth percentage is less than or equal to -30% or greater than or equal to 30%, the validation fails for that row.
    3. Otherwise, the validation passes for that row.

    The results of this validation are stored in a new column in the DataFrame with a prefix like "QUERY_03_" followed by the name of the first column in the pair. Additionally, the calculated growth percentage for each row is also stored in a separate column with a prefix "QUERY_03_GROWTH".

    Parameters:
    - df: DataFrame containing the data.
    - column_pairs: List of tuples, where each tuple contains a pair of column names. The growth is calculated using these columns.

    Returns:
    - DataFrame with added columns for validation results and growth calculations.
    - Dictionary containing the growth percentage for each column pair.

    Note:
    The function is designed to handle multiple column pairs, allowing for batch processing of growth validations.
    """
    
    growths = {}  # Create an empty dictionary to store the growth calculations
    for col1, col2 in column_pairs:  # Iterate over each column pair
        growth_col = f'QUERY_03_GROWTH{col1}'  # Create the name for the growth column
        result_col = f'QUERY_03_{col1}'  # Create the name for the result column
        growth = ((df[col1] / df[col2]) - 1) * 100  # Calculate the growth
        growths[growth_col] = growth  # Store the growth calculation in the dictionary
        df[result_col] = growth.apply(lambda x: 0 if x <= -30 or x >= 30 else 1)  # Add the result column to the DataFrame with pass/fail values based on the growth calculation
    return df, growths  # Return the updated DataFrame and the dictionary of growth calculations

df, growths_01 = query_03(df, column_pairs_01)
df, growths_02 = query_03(df, column_pairs_02)
df, growths_03 = query_03(df, column_pairs_03)

# -----------------------------------
# Query 4
# -----------------------------------

def generate_lists_04(df, year, quarter):
    def get_month_quarter(quarter):
        if quarter == 1:
            return ['01', '02', '03']
        elif quarter == 2:
            return ['04', '05', '06']
        elif quarter == 3:
            return ['07', '08', '09']
        elif quarter == 4:
            return ['10', '11', '12']
        else:
            return ['No quarter found']

    month_quarter = get_month_quarter(quarter)
    past_quarter = quarter - 1
    if past_quarter < 1:
        past_quarter = 4  # Wrap around to the fourth quarter if current quarter is 1
    past_month_quarter = get_month_quarter(past_quarter)  # Get month range for previous quarter

    A01_prefix = past_month_quarter[-1]
    A1_prefix, A2_prefix, A3_prefix = month_quarter
    C1_prefix, C2_prefix, C3_prefix = month_quarter
    X1_prefix, X2_prefix, X3_prefix = month_quarter

    lists = [[] for _ in range(10)]
    for i in range(1, 11):
        past_year = year - 1 if quarter == 1 else year  # Subtract 1 from the year only if the quarter is 1
        A01_col = f'A{past_year}{past_quarter}{A01_prefix}_F13{i:02d}'
        A1_col = f'A{year}{quarter}{A1_prefix}_F01{i:02d}'
        A2_col = f'A{year}{quarter}{A2_prefix}_F07{i:02d}'
        A3_col = f'A{year}{quarter}{A3_prefix}_F13{i:02d}'
        C1_col = f'C{year}{quarter}{C1_prefix}_F03{i:02d}'
        C2_col = f'C{year}{quarter}{C2_prefix}_F09{i:02d}'
        C3_col = f'C{year}{quarter}{C3_prefix}_F15{i:02d}'
        X1_col = f'X{year}{quarter}{X1_prefix}_F62{i:02d}'
        X2_col = f'X{year}{quarter}{X2_prefix}_F63{i:02d}'
        X3_col = f'X{year}{quarter}{X3_prefix}_F64{i:02d}'

        columns = [A01_col, A1_col, A2_col, A3_col, C1_col, C2_col, C3_col, X1_col, X2_col, X3_col]
        for j, col in enumerate(columns):
            lists[j].append(col)

    return tuple(lists)

lists_tuple = generate_lists_04(df, year, quarter)
A01_lst, A1_lst, A2_lst, A3_lst, C1_lst, C2_lst, C3_lst, X1_lst, X2_lst, X3_lst = lists_tuple

def query_04(df, A01_lst, A_lst, C_lst, X_lst):
    
    """
    This function is applied to a DataFrame to validate the relationship between:
    - the past month's total number of employees (A01),
    - the current month's total number of employees (A),
    - the current month's hiring count (C),
    - and the current month's total termination count (X).

    Logic:
    1. If the computed total (XYZ) is greater than the past month's employee count (A01 value):
        - The sum of the computed total (XYZ) and the current month's hiring count (C column) should be equal to the current month's total number of employees (A column).
    2. If the computed total (XYZ) is less than the past month's employee count (A01 value):
        - The difference between the computed total (XYZ) and the current month's total termination count (X column) should be equal to the current month's total number of employees (A column).
    3. If the computed total (XYZ) is equal to the past month's employee count (A01 value):
        - The current month's total number of employees (A column) remains unchanged.

    The results of this validation are stored in a new column in the DataFrame with a prefix like "QUERY_04_" followed by the A column name, where each value indicates whether the condition holds (1 for True, 0 for False).

    Parameters:
    - df: DataFrame containing the data.
    - A01_lst: List of columns representing the past month's employee count.
    - A_lst: List of columns representing the current month's employee count.
    - C_lst: List of columns representing the current month's hiring count.
    - X_lst: List of columns representing the current month's total termination count.

    Returns:
    - DataFrame with added columns for validation results.
    """

    for A01, A, C, X in zip(A01_lst, A_lst, C_lst, X_lst):
        if all(col in df.columns for col in [A01, A, C, X]):  # Check if required columns are present
            
            XYZ = df[A] + (df[A01] + df[C] - df[X])
            
            result_positive = ((XYZ > df[A01]) & (XYZ + df[C] == df[A])).astype(int)
            result_negative = ((XYZ < df[A01]) & (XYZ - df[X] == df[A])).astype(int)
            result_equal = (XYZ == df[A01]).astype(int)
            
            df[f"QUERY_04_{A}"] = result_positive + result_negative + result_equal
            
            
#           Comment/Uncomment this section to enable/disable print statements
            print(f"Processing columns: {A01}, {A}, {C}, {X}")
            print(f"Computed XYZ for {A}: {XYZ}")
            print(f"Validation result for {A}: {df[f'QUERY_04_{A}'].values}")
    
            
    return df

df = query_04(df, A01_lst, A1_lst, C1_lst, X1_lst)
df = query_04(df, A1_lst, A2_lst, C2_lst, X2_lst)
df = query_04(df, A2_lst, A3_lst, C3_lst, X3_lst)

# -----------------------------------
# Query 5
# -----------------------------------

def generate_lists_05(df, year, quarter,
                      A01_prefix=None, B01_prefix=None,
                      A1_prefix=None, A2_prefix=None, A3_prefix=None,
                      B1_prefix=None, B2_prefix=None, B3_prefix=None,
                      C1_prefix=None, C2_prefix=None, C3_prefix=None,
                      X1_prefix=None, X2_prefix=None, X3_prefix=None):

    def get_month_quarter(quarter):
        if quarter == 1:
            return ['01', '02', '03']
        elif quarter == 2:
            return ['04', '05', '06']
        elif quarter == 3:
            return ['07', '08', '09']
        elif quarter == 4:
            return ['10', '11', '12']
        else:
            return ['No quarter found']

    month_quarter = get_month_quarter(quarter)
    
    past_quarter = quarter - 1
    if past_quarter < 1:
        past_quarter = 4  # Wrap around to the fourth quarter if current quarter is 
        
    past_month_quarter = get_month_quarter(past_quarter)  # Get month range for previous quarter

    if A01_prefix is None:
        A01_prefix = past_month_quarter[-1]
    if B01_prefix is None:
        B01_prefix = past_month_quarter[-1]
    if A1_prefix is None:
        A1_prefix = month_quarter[0]
    if A2_prefix is None:
        A2_prefix = month_quarter[1]
    if A3_prefix is None:
        A3_prefix = month_quarter[2]
    if B1_prefix is None:
        B1_prefix = month_quarter[0]
    if B2_prefix is None:
        B2_prefix = month_quarter[1]
    if B3_prefix is None:
        B3_prefix = month_quarter[2]
    if C1_prefix is None:
        C1_prefix = month_quarter[0]
    if C2_prefix is None:
        C2_prefix = month_quarter[1]
    if C3_prefix is None:
        C3_prefix = month_quarter[2]
    if X1_prefix is None:
        X1_prefix = month_quarter[0]
    if X2_prefix is None:
        X2_prefix = month_quarter[1]
    if X3_prefix is None:
        X3_prefix = month_quarter[2]

    A01_lst = []
    B01_lst = []
    A1_lst = []
    A2_lst = []
    A3_lst = []
    B1_lst = []
    B2_lst = []
    B3_lst = []
    C1_lst = []
    C2_lst = []
    C3_lst = []
    X1_lst = []
    X2_lst = []
    X3_lst = []

    for i in range(1, 11):
        past_year = year - 1 if quarter == 1 else year  # Subtract 1 from the year only if the quarter is 1
        A01_col = f'A{past_year}{past_quarter}{A01_prefix}_F13{i:02d}'
        B01_col = f'B{past_year}{past_quarter}{B01_prefix}_F14{i:02d}'
        A1_col = f'A{year}{quarter}{A1_prefix}_F01{i:02d}'
        A2_col = f'A{year}{quarter}{A2_prefix}_F07{i:02d}'
        A3_col = f'A{year}{quarter}{A3_prefix}_F13{i:02d}'
        B1_col = f'B{year}{quarter}{B1_prefix}_F02{i:02d}'
        B2_col = f'B{year}{quarter}{B2_prefix}_F08{i:02d}'
        B3_col = f'B{year}{quarter}{B3_prefix}_F14{i:02d}'
        C1_col = f'C{year}{quarter}{C1_prefix}_F03{i:02d}'
        C2_col = f'C{year}{quarter}{C2_prefix}_F09{i:02d}'
        C3_col = f'C{year}{quarter}{C3_prefix}_F15{i:02d}'
        X1_col = f'X{year}{quarter}{X1_prefix}_F62{i:02d}'
        X2_col = f'X{year}{quarter}{X2_prefix}_F63{i:02d}'
        X3_col = f'X{year}{quarter}{X3_prefix}_F64{i:02d}'

        A01_lst.append(A01_col)
        B01_lst.append(B01_col)
        A1_lst.append(A1_col)
        A2_lst.append(A2_col)
        A3_lst.append(A3_col)
        B1_lst.append(B1_col)
        B2_lst.append(B2_col)
        B3_lst.append(B3_col)
        C1_lst.append(C1_col)
        C2_lst.append(C2_col)
        C3_lst.append(C3_col)
        X1_lst.append(X1_col)
        X2_lst.append(X2_col)
        X3_lst.append(X3_col)

    return A01_lst, B01_lst, A1_lst, A2_lst, A3_lst, B1_lst, B2_lst, B3_lst, C1_lst, C2_lst, C3_lst, X1_lst, X2_lst, X3_lst

lists_tuple = generate_lists_05(df, year, quarter)
A01_lst, B01_lst, A1_lst, A2_lst, A3_lst, B1_lst, B2_lst, B3_lst, C1_lst, C2_lst, C3_lst, X1_lst, X2_lst, X3_lst = lists_tuple      

def query_05(row, 
             A01_lst, B01_lst,
             A1_lst, A2_lst, A3_lst,
             B1_lst, B2_lst, B3_lst,
             C1_lst, C2_lst, C3_lst,
             X1_lst, X2_lst, X3_lst):
    
    """
    This function is applied row-wise to a DataFrame to validate the relationship
    between the total number of employees (A), the total number of vacant positions (B),
    the total number of employees in training (C), and the number of temporary staff (X).

    Logic:
    1. For the previous quarter:
        - Calculate the number of vacant positions for the current month: (B from previous quarter - C) + X.
    2. For the current quarter:
        - For each month in the quarter, compare the calculated number of vacant positions with the given number of vacant positions in the B column.
        - If the calculated number of vacant positions matches the given number in the B column, the check passes.
        - If the calculated number of vacant positions is greater than the number of employees in the A column for that month, the check fails.

    The results of this validation are stored in a dictionary with keys in the format `QUERY_05_{a_col}`,
    where each value indicates whether the condition holds (1 for True, 0 for False).

    Parameters:
    - row : pandas Series
        The current row of the DataFrame being processed.
    - Lists for each category (A, B, C, X) corresponding to each month of the quarter.

    Returns:
    - pd.Series
        Series containing the results of the validation for the row.
    """

    results = {}  # Create an empty dictionary to store the results
    
    for i in range(10):  # Iterate over the range of 10
        # Extract column names from generated lists for each category
        columns = {
            'A01': A01_lst[i],
            'B01': B01_lst[i],
            'A1': A1_lst[i],
            'A2': A2_lst[i],
            'A3': A3_lst[i],
            'B1': B1_lst[i],
            'B2': B2_lst[i],
            'B3': B3_lst[i],
            'C1': C1_lst[i],
            'C2': C2_lst[i],
            'C3': C3_lst[i],
            'X1': X1_lst[i],
            'X2': X2_lst[i],
            'X3': X3_lst[i]
        }
        
        # Printing the columns being processed
#         print(f"Processing columns: {columns['A01']}, {columns['A1']}, {columns['C1']}, {columns['X1']}")
#         print(f"Processing columns: {columns['B1']}, {columns['A2']}, {columns['C2']}, {columns['X2']}")
#         print(f"Processing columns: {columns['B2']}, {columns['A3']}, {columns['C3']}, {columns['X3']}")

        # Calculations for the three categories
        vacancy_present_01 = (row[columns['B01']] - row[columns['C1']]) + row[columns['X1']]
        vacancy_present_02 = (row[columns['B1']] - row[columns['C2']]) + row[columns['X2']]
        vacancy_present_03 = (row[columns['B2']] - row[columns['C3']]) + row[columns['X3']]

        # Validation for the first category
        if vacancy_present_01 == row[columns['B1']]:
            results[f'QUERY_05_{columns["A1"]}'] = 0
        else:
            results[f'QUERY_05_{columns["A1"]}'] = 1

        # Validation for the second category
        if vacancy_present_02 == row[columns['B2']]:
            results[f'QUERY_05_{columns["A2"]}'] = 0 if vacancy_present_02 > row[columns['A2']] else 1
        else:
            results[f'QUERY_05_{columns["A2"]}'] = 1
        
        # Validation for the third category
        if vacancy_present_03 == row[columns['B3']]:
            results[f'QUERY_05_{columns["A3"]}'] = 0 if vacancy_present_03 > row[columns['A3']] else 1
        else:
            results[f'QUERY_05_{columns["A3"]}'] = 1
        
    return pd.Series(results)

# Apply the combined function
results_vacant = df.apply(lambda row: query_05(row,
                                               A01_lst, B01_lst,
                                               A1_lst, A2_lst, A3_lst,
                                               B1_lst, B2_lst, B3_lst,
                                               C1_lst, C2_lst, C3_lst,
                                               X1_lst, X2_lst, X3_lst), axis=1)

# Assign the calculated values to the present DataFrame
df[results_vacant.columns] = results_vacant

# -----------------------------------
# Query 6
# -----------------------------------

def generate_lists_06(df, year, quarter,
                   A1_prefix=None, A2_prefix=None, A3_prefix=None,
                   G1_prefix=None, G2_prefix=None, G3_prefix=None,
                   H1_prefix=None, H2_prefix=None, H3_prefix=None):

    def get_month_quarter(quarter):
        if quarter == 1:
            return ['01', '02', '03']
        elif quarter == 2:
            return ['04', '05', '06']
        elif quarter == 3:
            return ['07', '08', '09']
        elif quarter == 4:
            return ['10', '11', '12']
        else:
            return ['No quarter found']

    month_quarter = get_month_quarter(quarter)

    if A1_prefix is None:
        A1_prefix = month_quarter[0]
    if A2_prefix is None:
        A2_prefix = month_quarter[1]
    if A3_prefix is None:
        A3_prefix = month_quarter[2]
    if G1_prefix is None:
        G1_prefix = month_quarter[0]
    if G2_prefix is None:
        G2_prefix = month_quarter[1]
    if G3_prefix is None:
        G3_prefix = month_quarter[2]
    if H1_prefix is None:
        H1_prefix = month_quarter[0]
    if H2_prefix is None:
        H2_prefix = month_quarter[1]
    if H3_prefix is None:
        H3_prefix = month_quarter[2]

    A1_lst = []
    A2_lst = []
    A3_lst = []
    G1_lst = []
    G2_lst = []
    G3_lst = []
    H1_lst = []
    H2_lst = []
    H3_lst = []

    for i in range(1, 11):
        A1_col = f'A{year}{quarter}{A1_prefix}_F01{i:02d}'
        A2_col = f'A{year}{quarter}{A2_prefix}_F07{i:02d}'
        A3_col = f'A{year}{quarter}{A3_prefix}_F13{i:02d}'
        G1_col = f'G{year}{quarter}{G1_prefix}_F23{i:02d}'
        G2_col = f'G{year}{quarter}{G2_prefix}_F29{i:02d}'
        G3_col = f'G{year}{quarter}{G3_prefix}_F35{i:02d}'
        H1_col = f'H{year}{quarter}{H1_prefix}_F24{i:02d}'
        H2_col = f'H{year}{quarter}{H2_prefix}_F30{i:02d}'
        H3_col = f'H{year}{quarter}{H3_prefix}_F36{i:02d}'

        A1_lst.append(A1_col)
        A2_lst.append(A2_col)
        A3_lst.append(A3_col)
        G1_lst.append(G1_col)
        G2_lst.append(G2_col)
        G3_lst.append(G3_col)
        H1_lst.append(H1_col)
        H2_lst.append(H2_col)
        H3_lst.append(H3_col)

    return A1_lst, A2_lst, A3_lst, G1_lst, G2_lst, G3_lst, H1_lst, H2_lst, H3_lst

lists_tuple = generate_lists_06(df, year, quarter)
A1_lst, A2_lst, A3_lst, G1_lst, G2_lst, G3_lst, H1_lst, H2_lst, H3_lst = lists_tuple

def validation_06(df, A_lsts, G_lsts, H_lsts):
    """
    This function is applied row-wise to a DataFrame to validate the relationship
    between the total number of employees (A), the total number of employees in standard jobs (G), 
    and the total number of employees in non-standard jobs (H).

    Logic:
    1. If there are no employees (value in A column is NaN or 0):
        - Both the number of employees in standard jobs (G column) and non-standard jobs (H column) should be NaN or 0.
    2. If there are employees present (value in A column is greater than 0):
        - Both the number of employees in standard jobs (G column) and non-standard jobs (H column) should not be NaN.
    3. The sum of the number of employees in standard jobs (G column) and non-standard jobs (H column) should be close to 
       the total number of employees (A column) without exceeding it.

    The results of this validation are stored in three new columns for each A column. The new columns are named 
    `QUERY_06_01_{a_col}`, `QUERY_06_02_{a_col}`, and `QUERY_06_03_{a_col}` respectively, 
    where each value indicates whether the condition holds (1 for True, 0 for False).

    Parameters:
    - df : pandas DataFrame
        The input DataFrame on which validation is to be applied.
    - A_lsts : list of lists
        Lists containing the names of the 'A' columns to be validated.
    - G_lsts : list of lists
        Lists containing the names of the 'G' columns to be validated.
    - H_lsts : list of lists
        Lists containing the names of the 'H' columns to be validated.

    Returns:
    - pd.DataFrame
        DataFrame with the results of the validation added as new columns.
    """

    for idx, A_cols in enumerate(A_lsts):
        A_lst = A_cols
        G_lst = G_lsts[idx]
        H_lst = H_lsts[idx]

        for i, row in df.iterrows():
            for A, G, H in zip(A_lst, G_lst, H_lst):
                # Define result column names for each condition
                result_col_1 = "QUERY_06_01_" + A
                result_col_2 = "QUERY_06_02_" + A
                result_col_3 = "QUERY_06_03_" + A

                # Condition 1: Handling NaN or 0 in 'A' columns
                if (pd.isna(row[A]) or row[A] == 0) and (pd.isna(row[G]) or row[G] == 0) and (pd.isna(row[H]) or row[H] == 0):
                    df.at[i, result_col_1] = 1  # Pass
                else:
                    df.at[i, result_col_1] = 0  # Fail

                # Condition 2: Handling non-NaN values in 'A' columns when 'G' or 'H' columns have NaN
                if row[A] > 0 and (pd.isna(row[G]) or pd.isna(row[H])):
                    df.at[i, result_col_2] = 0  # Fail
                else:
                    df.at[i, result_col_2] = 1  # Pass

                # Condition 3: Checking the difference between 'A' and the sum of 'G' and 'H'
                if pd.notnull(row[A]) and (pd.notnull(row[G]) and pd.notnull(row[H])):
                    diff = abs(row[A] - (row[G] + row[H]))
                    if diff <= row[A] or diff <= 0.9 * row[A]:
                        df.at[i, result_col_3] = 1  # Pass
                    else:
                        df.at[i, result_col_3] = 0  # Fail
                else:
                    df.at[i, result_col_3] = 0  # Fail

    return df

validation_06(df, [A1_lst, A2_lst, A3_lst], [G1_lst, G2_lst, G3_lst], [H1_lst, H2_lst, H3_lst])

# -----------------------------------
# Query 7
# -----------------------------------

def generate_lists_07(df, year, quarter,
                      A_prefix='07', I_prefix='07', J_prefix='07'):

    def get_month_quarter(quarter):
        if quarter == 1:
            return ['01', '02', '03']
        elif quarter == 2:
            return ['04', '05', '06']
        elif quarter == 3:
            return ['07', '08', '09']
        elif quarter == 4:
            return ['10', '11', '12']
        else:
            return ['No quarter found']

    month_quarter = get_month_quarter(quarter)

    if A_prefix is None:
        A_prefix = month_quarter[0]
    if I_prefix is None:
        I_prefix = month_quarter[1]
    if J_prefix is None:
        J_prefix = month_quarter[2]
    
    A_lst = []
    I_lst = []
    J_lst = []
    
    for i in range(1, 11):
        A_col = f'A{year}{quarter}{A_prefix}_F01{i:02d}'
        I_col = f'I{year}{quarter}{I_prefix}_F25{i:02d}'
        J_col = f'J{year}{quarter}{J_prefix}_F26{i:02d}'

        A_lst.append(A_col)
        I_lst.append(I_col)
        J_lst.append(J_col)

    return A_lst, I_lst, J_lst

lists_tuple = generate_lists_07(df, year, quarter)
A_lst, I_lst, J_lst = lists_tuple

def query_07(row):
    """
    This function is applied row-wise to a DataFrame to validate the relationship
    between the number of employees (A), the total working days per month (I),
    and the working hours per day (J).

    Logic:
    1. If there are no employees (value in A column is 0):
        - Both the total working days (I column) and working hours per day (J column) should be 0.
    2. If there are employees present (value in A column is not 0):
        - At least one of the total working days (I column) or working hours per day (J column) should be non-zero.

    The results of this validation are stored in a dictionary with keys in the format `QUERY_07_{a_col}_{i_col}`,
    where each value indicates whether the condition holds (1 for True, 0 for False).

    Parameters:
    - row : pandas Series
        The current row of the DataFrame being processed.

    Returns:
    - pd.Series
        Series containing the results of the validation for the row.
    """
    
    result = {}
    for a_col, i_col, j_col in zip(A_lst, I_lst, J_lst):
        print(f"Processing columns: {a_col}, {i_col}, {j_col}")  # Indicate the columns being processed.
        if row[a_col] == 0:
            result[f'QUERY_07_{a_col}_{i_col}'] = int(row[i_col] == 0 and row[j_col] == 0)
        else:
            result[f'QUERY_07_{a_col}_{i_col}'] = int(row[i_col] != 0 or row[j_col] != 0)
    return pd.Series(result)

df = df.join(df.apply(query_07, axis=1))

# -----------------------------------
# Query 8
# -----------------------------------

def generate_lists_08(year, quarter,
                      prefix_a_01='07', prefix_o_01='07', prefix_k_01='07',
                      prefix_a_02='08', prefix_o_02='08', prefix_k_02='08',
                      prefix_a_03='09', prefix_o_03='09', prefix_k_03='09'):

    def get_month_quarter(quarter):
        if quarter == 1:
            return ['01', '02', '03']
        elif quarter == 2:
            return ['04', '05', '06']
        elif quarter == 3:
            return ['07', '08', '09']
        elif quarter == 4:
            return ['10', '11', '12']
        else:
            return ['No quarter found']

    month_quarter = get_month_quarter(quarter)
    
    a_01_col = f'A{year}{quarter}{prefix_a_01}_F0110'
    o_01_col = f'O{year}{quarter}{prefix_o_01}_F4410'
    k_01_col = f'K{year}{quarter}{prefix_k_01}_F2810'
    a_02_col = f'A{year}{quarter}{prefix_a_02}_F0710'
    o_02_col = f'O{year}{quarter}{prefix_o_02}_F5010'
    k_02_col = f'K{year}{quarter}{prefix_k_02}_F3410'
    a_03_col = f'A{year}{quarter}{prefix_a_03}_F1310'
    o_03_col = f'O{year}{quarter}{prefix_o_03}_F5610'
    k_03_col = f'K{year}{quarter}{prefix_k_03}_F4010'

    column_pairs = [(a_01_col, o_01_col, k_01_col), 
                    (a_02_col, o_02_col, k_02_col), 
                    (a_03_col, o_03_col, k_03_col)]

    return column_pairs

column_pairs = generate_lists_08(year, quarter)

def query_08(df, column_pairs):
    """
    This function is applied to a DataFrame to validate the relationship among 
    the number of employees (A), the overtime pay (O), and the total hours worked overtime (K).

    Logic:
    1. If the number of employees (value in A column) has a value (not null and not zero):
        - If the overtime pay (O column) has a value (not null and not zero), the total hours worked overtime (K column) can have any value.
        - If the overtime pay (O column) doesn't have a value (either null or zero), then the total hours worked overtime (K column) must also not have a value (either null or zero).
    2. If the number of employees (value in A column) doesn't have a value (either null or zero):
        - Both the overtime pay (O column) and the total hours worked overtime (K column) should not have values (either null or zero).

    The results of this validation are stored in new columns in the DataFrame, with names in the format `QUERY_08_{a_col}_{o_col}_{k_col}`,
    where each value indicates whether the condition holds (1 for True, 0 for False).

    Parameters:
    - df (pd.DataFrame): The input dataframe on which the validation is applied.
    - column_pairs (list of tuple): A list containing triples of column names. The first column in each triple corresponds to column A, the second to column O, and the third to column K.

    Returns:
    - pd.DataFrame: A dataframe with new columns added for the validation results.
    """

    for col_A, col_O, col_K in column_pairs:
        
        # Condition for column A having a value
        A_has_value = df[col_A].notnull() & (df[col_A] != 0)
        
        # Sub-condition for column O having a value when A has value
        O_has_value_when_A = A_has_value & df[col_O].notnull() & (df[col_O] != 0)
        
        # Sub-condition for column O not having a value when A has value
        O_no_value_when_A = A_has_value & (df[col_O].isnull() | (df[col_O] == 0))
        
        # Condition for column K having no value when column O doesn't have a value
        K_no_value_when_O_no_value = O_no_value_when_A & (df[col_K].isnull() | (df[col_K] == 0))
        
        # Condition for column A not having a value
        A_no_value = df[col_A].isnull() | (df[col_A] == 0)
        
        # Both columns O and K should not have values when A doesn't have a value
        O_and_K_no_value_when_A_no_value = A_no_value & (df[col_O].isnull() | (df[col_O] == 0)) & (df[col_K].isnull() | (df[col_K] == 0))

        # Combine all conditions
        result = O_has_value_when_A | K_no_value_when_O_no_value | O_and_K_no_value_when_A_no_value
        
        # Convert the boolean result to integer (1 for pass, 0 for fail)
        df[f"QUERY_08_{col_A}_{col_O}_{col_K}"] = result.astype(int)
        
    return df

# Apply the validation
df = query_08(df, column_pairs)

# -----------------------------------
# Query 9
# -----------------------------------

def generate_lists_09(df,
                      A1_prefix=None, A2_prefix=None, A3_prefix=None,
                      L1_prefix=None, L2_prefix=None, L3_prefix=None,
                      M1_prefix=None, M2_prefix=None, M3_prefix=None,
                      N1_prefix=None, N2_prefix=None, N3_prefix=None,
                      O1_prefix=None, O2_prefix=None, O3_prefix=None,
                      P1_prefix=None, P2_prefix=None, P3_prefix=None,
                      Q1_prefix=None, Q2_prefix=None, Q3_prefix=None,
                      R1_prefix=None, R2_prefix=None, R3_prefix=None):

    def get_month_quarter(quarter):
        if quarter == 1:
            return ['01', '02', '03']
        elif quarter == 2:
            return ['04', '05', '06']
        elif quarter == 3:
            return ['07', '08', '09']
        elif quarter == 4:
            return ['10', '11', '12']
        else:
            return ['No quarter found']

    month_quarter = get_month_quarter(quarter)

    if A1_prefix is None:
        A1_prefix = month_quarter[0]
    if A2_prefix is None:
        A2_prefix = month_quarter[1]
    if A3_prefix is None:
        A3_prefix = month_quarter[2]
    if L1_prefix is None:
        L1_prefix = month_quarter[0]
    if L2_prefix is None:
        L2_prefix = month_quarter[1]
    if L3_prefix is None:
        L3_prefix = month_quarter[2]
    if M1_prefix is None:
        M1_prefix = month_quarter[0]
    if M2_prefix is None:
        M2_prefix = month_quarter[1]
    if M3_prefix is None:
        M3_prefix = month_quarter[2]
    if N1_prefix is None:
        N1_prefix = month_quarter[0]
    if N2_prefix is None:
        N2_prefix = month_quarter[1]
    if N3_prefix is None:
        N3_prefix = month_quarter[2]
    if O1_prefix is None:
        O1_prefix = month_quarter[0]
    if O2_prefix is None:
        O2_prefix = month_quarter[1]
    if O3_prefix is None:
        O3_prefix = month_quarter[2]
    if P1_prefix is None:
        P1_prefix = month_quarter[0]
    if P2_prefix is None:
        P2_prefix = month_quarter[1]
    if P3_prefix is None:
        P3_prefix = month_quarter[2]
    if Q1_prefix is None:
        Q1_prefix = month_quarter[0]
    if Q2_prefix is None:
        Q2_prefix = month_quarter[1]
    if Q3_prefix is None:
        Q3_prefix = month_quarter[2]
    if R1_prefix is None:
        R1_prefix = month_quarter[0]
    if R2_prefix is None:
        R2_prefix = month_quarter[1]
    if R3_prefix is None:
        R3_prefix = month_quarter[2]

    A1_lst = []
    A2_lst = []
    A3_lst = []
    L1_lst = []
    L2_lst = []
    L3_lst = []
    M1_lst = []
    M2_lst = []
    M3_lst = []
    N1_lst = []
    N2_lst = []
    N3_lst = []
    O1_lst = []
    O2_lst = []
    O3_lst = []
    P1_lst = []
    P2_lst = []
    P3_lst = []
    Q1_lst = []
    Q2_lst = []
    Q3_lst = []
    R1_lst = []
    R2_lst = []
    R3_lst = []

    for i in range(1, 11):
        A1_col = f'A{year}{quarter}{A1_prefix}_F01{i:02d}'
        A2_col = f'A{year}{quarter}{A2_prefix}_F07{i:02d}'
        A3_col = f'A{year}{quarter}{A3_prefix}_F13{i:02d}'
        L1_col = f'L{year}{quarter}{L1_prefix}_F41{i:02d}'
        L2_col = f'L{year}{quarter}{L2_prefix}_F47{i:02d}'
        L3_col = f'L{year}{quarter}{L3_prefix}_F53{i:02d}'
        M1_col = f'M{year}{quarter}{M1_prefix}_F42{i:02d}'
        M2_col = f'M{year}{quarter}{M2_prefix}_F48{i:02d}'
        M3_col = f'M{year}{quarter}{M3_prefix}_F54{i:02d}'
        N1_col = f'N{year}{quarter}{N1_prefix}_F43{i:02d}'
        N2_col = f'N{year}{quarter}{N2_prefix}_F49{i:02d}'
        N3_col = f'N{year}{quarter}{N3_prefix}_F55{i:02d}'
        O1_col = f'O{year}{quarter}{O1_prefix}_F44{i:02d}'
        O2_col = f'O{year}{quarter}{O2_prefix}_F50{i:02d}'
        O3_col = f'O{year}{quarter}{O3_prefix}_F56{i:02d}'
        P1_col = f'P{year}{quarter}{P1_prefix}_F59{i:02d}'
        P2_col = f'P{year}{quarter}{P2_prefix}_F60{i:02d}'
        P3_col = f'P{year}{quarter}{P3_prefix}_F61{i:02d}'
        Q1_col = f'Q{year}{quarter}{P1_prefix}_F45{i:02d}'
        Q2_col = f'Q{year}{quarter}{P2_prefix}_F51{i:02d}'
        Q3_col = f'Q{year}{quarter}{P3_prefix}_F57{i:02d}'
        R1_col = f'R{year}{quarter}{P1_prefix}_F46{i:02d}'
        R2_col = f'R{year}{quarter}{P2_prefix}_F52{i:02d}'
        R3_col = f'R{year}{quarter}{P3_prefix}_F58{i:02d}'

        A1_lst.append(A1_col)
        A2_lst.append(A2_col)
        A3_lst.append(A3_col)
        L1_lst.append(L1_col)
        L2_lst.append(L2_col)
        L3_lst.append(L3_col)
        M1_lst.append(M1_col)
        M2_lst.append(M2_col)
        M3_lst.append(M3_col)
        N1_lst.append(N1_col)
        N2_lst.append(N2_col)
        N3_lst.append(N3_col)
        O1_lst.append(O1_col)
        O2_lst.append(O2_col)
        O3_lst.append(O3_col)
        P1_lst.append(P1_col)
        P2_lst.append(P2_col)
        P3_lst.append(P3_col)
        Q1_lst.append(Q1_col)
        Q2_lst.append(Q2_col)
        Q3_lst.append(Q3_col)
        R1_lst.append(R1_col)
        R2_lst.append(R2_col)
        R3_lst.append(R3_col)

    return A1_lst, A2_lst, A3_lst, L1_lst, L2_lst, L3_lst, M1_lst, M2_lst, M3_lst, N1_lst, N2_lst, N3_lst, O1_lst, O2_lst, O3_lst, P1_lst, P2_lst, P3_lst, Q1_lst, Q2_lst, Q3_lst, R1_lst, R2_lst, R3_lst                                 

year = 22
quarter = 3
lists_tuple = generate_lists_09(df)
A1_lst, A2_lst, A3_lst, L1_lst, L2_lst, L3_lst, M1_lst, M2_lst, M3_lst, N1_lst, N2_lst, N3_lst, O1_lst, O2_lst, O3_lst, P1_lst, P2_lst, P3_lst, Q1_lst, Q2_lst, Q3_lst, R1_lst, R2_lst, R3_lst = lists_tuple     

def query_09(df, generate_lists_09):
    """
    This function is applied to a DataFrame to validate the relationship 
    between the number of employees represented by columns under the 'A' naming convention 
    and the various payment types represented by columns under the 'L', 'M', 'N', 'O', and 'P' naming conventions.

    Logic:
    Validation 1:
    - If column 'A' has a value (indicating the number of employees):
        - At least one of the columns 'L', 'M', 'N', 'O', or 'P' must have a value, indicating some form of payment.
    - If column 'A' does not have a value (indicating no employees):
        - All of the columns 'L', 'M', 'N', 'O', and 'P' should also have no value, indicating no payments.
    The results of this validation are stored in columns named `QUERY_09_01_{main_column}`.

    Validation 2:
    - If column 'A' doesn't have a value (or is NaN) and column 'Q' has a value (indicating total salary and wages):
        - The result is considered a fail as there shouldn't be a total salary and wages without employees.
    The results of this validation are stored in columns named `QUERY_09_02_{main_column}`.

    Parameters:
    - df : pandas DataFrame
        The DataFrame containing the data.
    - generate_lists_09 : function
        A function that generates the required column lists based on the naming convention.

    Returns:
    - DataFrame
        DataFrame with added columns for validation results.
    """

    # Get the column name lists
    A1_lst, A2_lst, A3_lst, L1_lst, L2_lst, L3_lst, M1_lst, M2_lst, M3_lst, N1_lst, N2_lst, N3_lst, O1_lst, O2_lst, O3_lst, P1_lst, P2_lst, P3_lst, Q1_lst, Q2_lst, Q3_lst, R1_lst, R2_lst, R3_lst = generate_lists_09(df)    
    
    # Loop through each main_column in A1_lst
    for main_column in A1_lst:
        A_has_value = df[main_column].notnull() & df[main_column].ne(0)
        
        # Extract the relevant L to P and Q columns based on the main_column's index
        index = A1_lst.index(main_column)
        column_lists = [L1_lst[index], M1_lst[index], N1_lst[index], O1_lst[index], P1_lst[index]]
        Q_column = Q1_lst[index]
        
        L_to_P_have_value = df[column_lists].ne(0).any(axis=1)
        Q_has_value = df[Q_column].notnull() & df[Q_column].ne(0)
        
        # Condition 1
        result_condition_1 = ((A_has_value & L_to_P_have_value) | (~A_has_value & ~L_to_P_have_value))
        df[f"QUERY_09_01_{main_column}"] = result_condition_1.astype(int)
        
        # Condition 2
        result_condition_2 = ~(~A_has_value & Q_has_value)
        df[f"QUERY_09_02_{main_column}"] = result_condition_2.astype(int)

    return df

df = query_09(df, generate_lists_09)

# -----------------------------------
# Query 10
# -----------------------------------

def generate_lists_10(df, year, quarter,
                      A01_prefix=None,
                      A1_prefix=None, A2_prefix=None, A3_prefix=None,
                      Q01_prefix=None,
                      Q1_prefix=None, Q2_prefix=None, Q3_prefix=None):

    def get_month_quarter(quarter):
        if quarter == 1:
            return ['01', '02', '03']
        elif quarter == 2:
            return ['04', '05', '06']
        elif quarter == 3:
            return ['07', '08', '09']
        elif quarter == 4:
            return ['10', '11', '12']
        else:
            return ['No quarter found']

    month_quarter = get_month_quarter(quarter)
    
    past_quarter = quarter - 1
    if past_quarter < 1:
        past_quarter = 4  # Wrap around to the fourth quarter if current quarter is 
        
    past_month_quarter = get_month_quarter(past_quarter)  # Get month range for previous quarter

    if A01_prefix is None:
        A01_prefix = past_month_quarter[-1]
    if A1_prefix is None:
        A1_prefix = month_quarter[0]
    if A2_prefix is None:
        A2_prefix = month_quarter[1]
    if A3_prefix is None:
        A3_prefix = month_quarter[2]
    if Q01_prefix is None:
        Q01_prefix = past_month_quarter[-1]
    if Q1_prefix is None:
        Q1_prefix = month_quarter[0]
    if Q2_prefix is None:
        Q2_prefix = month_quarter[1]
    if Q3_prefix is None:
        Q3_prefix = month_quarter[2]
        
    A01_lst = []
    A1_lst = []
    A2_lst = []
    A3_lst = []
    Q01_lst = []
    Q1_lst = []
    Q2_lst = []
    Q3_lst = []

    for i in range(1, 11):
        past_year = year - 1 if quarter == 1 else year  # Subtract 1 from the year only if the quarter is 1
        A01_col = f'A{past_year}{past_quarter}{A01_prefix}_F13{i:02d}'
        A1_col = f'A{year}{quarter}{A1_prefix}_F01{i:02d}'
        A2_col = f'A{year}{quarter}{A2_prefix}_F07{i:02d}'
        A3_col = f'A{year}{quarter}{A3_prefix}_F13{i:02d}'
        Q01_col = f'Q{past_year}{past_quarter}{Q01_prefix}_F57{i:02d}'
        Q1_col = f'Q{year}{quarter}{Q1_prefix}_F45{i:02d}'
        Q2_col = f'Q{year}{quarter}{Q2_prefix}_F51{i:02d}'
        Q3_col = f'Q{year}{quarter}{Q3_prefix}_F57{i:02d}'

        A01_lst.append(A01_col)
        A1_lst.append(A1_col)
        A2_lst.append(A2_col)
        A3_lst.append(A3_col)
        Q01_lst.append(Q01_col)
        Q1_lst.append(Q1_col)
        Q2_lst.append(Q2_col)
        Q3_lst.append(Q3_col)

    return A01_lst, A1_lst, A2_lst, A3_lst, Q01_lst, Q1_lst, Q2_lst, Q3_lst

lists_tuple = generate_lists_10(df, year, quarter)
A01_lst, A1_lst, A2_lst, A3_lst, Q01_lst, Q1_lst, Q2_lst, Q3_lst = lists_tuple      

def query_10(df, A01_lst, Q01_lst, A1_lst, A2_lst, A3_lst, Q1_lst, Q2_lst, Q3_lst):
    """
    Perform validation based on the SGTGU dataset's naming convention.
    
    Purpose:
    - Checks for significant changes in the average salary of employees across different months.
    - The average salary is calculated as the ratio of "Total salary & wages" (Q) to the number of "Employees" (A).
    - The growth rate between consecutive months is then calculated, and the result is validated against a threshold.

    Pass/Fail Condition:
    - A row passes if the calculated growth rate for any month falls within the range [-30%, 30%].
    - A row fails if the calculated growth rate for any month falls outside this range.
    
    Variables Used:
    A: Employees
    Q: Total salary & wages
    
    Parameters:
    - df: DataFrame containing the data.
    - generate_lists_10: Function to generate the required column lists.
    
    Returns:
    - DataFrame with added columns for validation results.
    """
    
    results = {}
    for i in range(10):
        # Calculate average Total Salary & Wages per Employee for each month
        avg_A01 = df[Q01_lst[i]] / df[A01_lst[i]]
        avg_A1 = df[Q1_lst[i]] / df[A1_lst[i]]
        avg_A2 = df[Q2_lst[i]] / df[A2_lst[i]]
        avg_A3 = df[Q3_lst[i]] / df[A3_lst[i]]

        # Calculate growth rate between consecutive months
        growth_A1 = ((avg_A1 - avg_A01) / avg_A01) * 100
        growth_A2 = ((avg_A2 - avg_A1) / avg_A1) * 100
        growth_A3 = ((avg_A3 - avg_A2) / avg_A2) * 100

        # Validate if the growth rate is within -30% to 30%
        results[A1_lst[i]] = (growth_A1 >= -30) & (growth_A1 <= 30)
        results[A2_lst[i]] = (growth_A2 >= -30) & (growth_A2 <= 30)
        results[A3_lst[i]] = (growth_A3 >= -30) & (growth_A3 <= 30)

    # Store the validation results in the DataFrame
    for column, result in results.items():
        df['QUERY_10_' + column] = result.astype(int)

    return df

lists_tuple = generate_lists_10(df, year, quarter)
A01_lst, A1_lst, A2_lst, A3_lst, Q01_lst, Q1_lst, Q2_lst, Q3_lst = lists_tuple
validity_results = query_10(df, A01_lst, Q01_lst, A1_lst, A2_lst, A3_lst, Q1_lst, Q2_lst, Q3_lst)

# -----------------------------------
# Query 11
# -----------------------------------

def generate_lists_11(df, prefix):

    def get_month_quarter(quarter):
        if quarter == 1:
            return ['01', '02', '03']
        elif quarter == 2:
            return ['04', '05', '06']
        elif quarter == 3:
            return ['07', '08', '09']
        elif quarter == 4:
            return ['10', '11', '12']
        else:
            return ['No quarter found']

    month_quarter = get_month_quarter(quarter)
    
    prev_quarter = quarter - 1 if quarter > 1 else 4
    prev_month_quarter = get_month_quarter(prev_quarter)

    first_month_past = prev_month_quarter[-3]
    second_month_past = prev_month_quarter[-2]
    last_month_past = prev_month_quarter[-1]
    first_month_present = month_quarter[0]
    second_month_present = month_quarter[1]
    last_month_present = month_quarter[2]

    column_lists = {}
    for i in range(1, 11):
        list_name = f"{prefix}{i:02d}_11_lst"
        column_list = [f"{prefix}{year}{prev_quarter}{first_month_past}_F01{i:02d}",
                       f"{prefix}{year}{prev_quarter}{second_month_past}_F07{i:02d}",
                       f"{prefix}{year}{prev_quarter}{last_month_past}_F13{i:02d}",
                       f"{prefix}{year}{quarter}{first_month_present}_F01{i:02d}",
                       f"{prefix}{year}{quarter}{second_month_present}_F07{i:02d}",
                       f"{prefix}{year}{quarter}{last_month_present}_F13{i:02d}"]
        column_lists[list_name] = column_list
    return column_lists

column_lists = generate_lists_11(df, 'A')

# Access the list "A02_11_lst"
A01_11_lst = column_lists['A01_11_lst']
A02_11_lst = column_lists['A02_11_lst']
A03_11_lst = column_lists['A03_11_lst']
A04_11_lst = column_lists['A04_11_lst']
A05_11_lst = column_lists['A05_11_lst']
A06_11_lst = column_lists['A06_11_lst']
A07_11_lst = column_lists['A07_11_lst']
A08_11_lst = column_lists['A08_11_lst']
A09_11_lst = column_lists['A09_11_lst']
A10_11_lst = column_lists['A10_11_lst']

def query_11(df, column_lists):
    """
    Perform validation based on the SGTGU dataset's naming convention.
    
    Purpose:
    - Checks for consistent changes across different categories for a company.
    - If all categories show no changes across all months, then the company fails the validation. Otherwise, it passes.

    Pass/Fail Condition:
    - A row passes if at least one category shows changes across any month.
    - A row fails if all categories show no changes across all months.
    
    Variables Used:
    A: Employees
    
    Parameters:
    - df: DataFrame containing the data.
    - generate_lists_11: Function to generate the required column lists.
    
    Returns:
    - DataFrame with added columns for validation results.
    """
    
    # Initialize the 'QUERY_11' column to 1 (indicating all companies pass by default).
    df['QUERY_11'] = 1
    
    # Iterate over each row (company) in the dataframe.
    for index, row in df.iterrows():
        has_change = False  # Variable to track if any change is found for the current company.
        
        # Iterate over each category list.
        for columns in column_lists.values():
            past_cols = columns[:3]
            present_cols = columns[3:]
            
            # Compare past and present data for the category.
            if not (row[past_cols].values == row[present_cols].values).all():
                has_change = True
                break
        
        # If no changes are found across all categories, mark the company as failing the validation.
        if not has_change:
            df.at[index, 'QUERY_11'] = 0

    return df

column_lists = generate_lists_11(df, 'A')
df_result = query_11(df, column_lists)

# Extract the original file name
original_file_name = os.path.basename(dataset_path)

output_file_path = ('C:/Users/aiman/Desktop/gh_konsistensi/output/sgtgu/')
suffix = '_new'

# Create the new file name by adding the suffix
new_file_name = original_file_name.replace('.csv', '') + suffix + '.csv'

# Save the DataFrame as CSV using the new file name
df.to_csv(os.path.join(output_file_path, new_file_name), index=False)